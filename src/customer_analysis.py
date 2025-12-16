#!/usr/bin/env python3
"""
Repeat customer purchase-pattern analysis builder.

Reads a multi-sheet Excel workbook with year tabs containing "Name of Customer" rows
and month columns, then produces an analysis workbook that is better suited for
discontinuous / lumpy repeat-customer behavior.

Primary outputs:
- Rolling concentration (TTM, 24M, 36M)
- Customer-level summary (tenure, gaps, reactivations, peak reliance, persistence)
- Top customer tables (lifetime, TTM, peak TTM share)

Usage:
  python src/generate_repeat_customer_analysis.py --input "4.0.5 Financial - Month-wise Customer Revenue.xlsx" \
      --output "Repeat_Customer_Purchase_Patterns_Starter_Analysis.xlsx"

Notes:
- Customer names are normalized (uppercase + whitespace collapse) for aggregation.
- The dataset is automatically trimmed to the last month with non-zero total revenue.
"""

from __future__ import annotations

import argparse
import math
import re
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd


MONTH_MAP: Dict[str, int] = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


def normalize_customer_name(name: str) -> str:
    """Normalize customer names for aggregation (case + whitespace)."""
    s = str(name).strip()
    s = re.sub(r"\s+", " ", s)
    return s.upper()


def _get_distinctive_name(name: str) -> str:
    """
    Extract the distinctive part of a company name (before generic suffixes).

    For "ALIGNED TECHNOLOGIES INC." returns "ALIGNED"
    For "SUN LIFE ASSURANCE COMPANY" returns "SUN LIFE ASSURANCE"
    """
    # Remove common business entity types and generic terms
    generic_terms = [
        'inc', 'incorporated', 'corp', 'corporation', 'ltd', 'limited',
        'llc', 'lp', 'ulc', 'plc', 'llp', 'company', 'co',
        'technologies', 'technology', 'services', 'group', 'partners',
        'canada', 'canadian', 'international', 'global',
    ]

    tokens = name.lower().split()

    # Keep tokens that aren't generic (but keep at least the first token)
    distinctive = []
    for i, token in enumerate(tokens):
        # Strip punctuation for comparison
        clean_token = token.strip('.,')
        # Always keep first token, then only keep non-generic tokens
        if i == 0 or clean_token not in generic_terms:
            distinctive.append(token)

    return ' '.join(distinctive) if distinctive else name.lower()


def _similarity_ratio(a: str, b: str) -> float:
    """Calculate similarity ratio between two strings (0.0 to 1.0)."""
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


def _strip_legal_suffix(name: str) -> str:
    """
    Remove common legal suffixes for comparison.

    Returns the name without suffixes like Inc., Corp., Ltd., etc.
    """
    suffixes = [
        r'\s+inc\.?$',
        r'\s+incorporated$',
        r'\s+corp\.?$',
        r'\s+corporation$',
        r'\s+ltd\.?$',
        r'\s+limited$',
        r'\s+llc$',
        r'\s+lp$',
        r'\s+ulc$',
        r'\s+plc$',
        r'\s+llp\.?$',
    ]

    result = name.lower()
    for suffix_pattern in suffixes:
        result = re.sub(suffix_pattern, '', result, flags=re.IGNORECASE)

    return result.strip()


@dataclass
class DuplicateMatch:
    """Represents a potential duplicate customer relationship."""
    customer1: str
    customer2: str
    similarity: float
    confidence: str  # HIGH, MEDIUM, LOW
    reason: str


def find_potential_duplicates(customers: List[str]) -> List[DuplicateMatch]:
    """
    Find potential duplicate customer names using multiple heuristics.

    Args:
        customers: List of normalized customer names

    Returns:
        List of DuplicateMatch objects for potential duplicates
    """
    matches: List[DuplicateMatch] = []
    customers_sorted = sorted(set(customers))

    for i, cust1 in enumerate(customers_sorted):
        for cust2 in customers_sorted[i + 1:]:
            # Calculate basic similarity
            sim = _similarity_ratio(cust1, cust2)

            # Get distinctive parts (without generic terms like "Technologies Inc.")
            dist1 = _get_distinctive_name(cust1)
            dist2 = _get_distinctive_name(cust2)
            dist_sim = _similarity_ratio(dist1, dist2)

            # Check various duplicate patterns
            confidence = None
            reason = None

            # Pattern 1: One name contains the other (exact substring)
            if cust1.lower() in cust2.lower() or cust2.lower() in cust1.lower():
                if sim > 0.85:  # Avoid spurious matches like "IBI GROUP" in "ARCADIS | IBI GROUP"
                    confidence = "HIGH"
                    reason = "One name contains the other"

            # Pattern 2: Only differ by legal suffix
            if confidence is None:
                stripped1 = _strip_legal_suffix(cust1)
                stripped2 = _strip_legal_suffix(cust2)
                if stripped1 == stripped2 and stripped1:
                    confidence = "HIGH"
                    reason = "Legal suffix variation only"

            # Pattern 3: Very high similarity (likely same company)
            # BUT: distinctive parts must also be similar to avoid false positives
            if confidence is None and sim > 0.90 and dist_sim > 0.60:
                confidence = "HIGH"
                reason = f"Very high similarity ({sim:.2f})"

            # Pattern 4: High similarity with common patterns
            # Require distinctive similarity >0.50 to avoid matching unrelated "XYZ Technologies Inc." companies
            if confidence is None and sim > 0.75 and dist_sim > 0.50:
                # Check for location/division suffixes
                location_pattern = r'\s+-\s+\d+\s+.*(?:st|street|ave|avenue|rd|road)\.?$'
                if re.search(location_pattern, cust1.lower()) or re.search(location_pattern, cust2.lower()):
                    confidence = "MEDIUM"
                    reason = "Location suffix variation"
                elif "(" in cust1 or "(" in cust2:
                    confidence = "MEDIUM"
                    reason = "Parenthetical addition"
                else:
                    confidence = "MEDIUM"
                    reason = f"High similarity ({sim:.2f})"

            # Pattern 5: Moderate similarity (possible duplicate, review needed)
            # Also require distinctive parts to be somewhat similar
            if confidence is None and sim > 0.70 and dist_sim > 0.50:
                confidence = "LOW"
                reason = f"Moderate similarity ({sim:.2f})"

            # Record match if we found one
            if confidence:
                matches.append(DuplicateMatch(
                    customer1=cust1,
                    customer2=cust2,
                    similarity=sim,
                    confidence=confidence,
                    reason=reason
                ))

    return matches


def _find_header_row(df_raw: pd.DataFrame, max_scan_rows: int = 60) -> Optional[int]:
    """Find the row index containing 'Name of Customer'."""
    scan_n = min(len(df_raw), max_scan_rows)
    for i in range(scan_n):
        row = df_raw.iloc[i].astype(str).str.strip().str.lower()
        if (row == "name of customer").any():
            return i
    return None


def _month_cols_from_headers(headers: Sequence[object]) -> List[str]:
    """Return the month columns present in the header row."""
    month_cols: List[str] = []
    for h in headers:
        key = str(h).strip().lower()
        if key in MONTH_MAP:
            month_cols.append(str(h))
    return month_cols


def parse_year_sheet(xlsx_path: Path, sheet_name: str) -> pd.DataFrame:
    """
    Parse a year sheet into long format:
      date, year, month, customer_raw, customer, revenue
    """
    df_raw = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=None)
    header_row = _find_header_row(df_raw)
    if header_row is None:
        raise ValueError(f"Could not find 'Name of Customer' header in sheet '{sheet_name}'.")

    headers = df_raw.iloc[header_row].tolist()
    headers = [
        (h if not (isinstance(h, float) and math.isnan(h)) else f"unnamed_{j}")
        for j, h in enumerate(headers)
    ]

    df = df_raw.iloc[header_row + 1 :].copy()
    df.columns = headers
    df = df.dropna(how="all")

    # Identify customer column
    cust_col = next((c for c in df.columns if str(c).strip().lower() == "name of customer"), df.columns[0])

    # Identify month columns
    month_cols: List[str] = []
    for c in df.columns:
        key = str(c).strip().lower()
        if key in MONTH_MAP:
            month_cols.append(c)

    if not month_cols:
        raise ValueError(f"No month columns recognized in sheet '{sheet_name}'.")

    df2 = df[[cust_col] + month_cols].copy()
    df2 = df2.rename(columns={cust_col: "customer_raw"})

    # Clean customer names and drop empty rows
    df2["customer_raw"] = df2["customer_raw"].astype(str).str.strip()
    df2.loc[df2["customer_raw"].str.lower().isin(["nan", ""]), "customer_raw"] = np.nan
    df2 = df2.dropna(subset=["customer_raw"])

    # Drop obvious totals rows
    df2 = df2[~df2["customer_raw"].str.lower().isin(["total", "totals", "grand total"])]

    # Coerce numeric revenue; treat "-", "–", "—" as 0
    for c in month_cols:
        # Replace dash characters and empty strings with NaN, then convert to numeric
        col_series = df2[c].astype(str)
        col_series = col_series.replace(["-", "–", "—", ""], np.nan)
        df2[c] = pd.to_numeric(col_series, errors="coerce").fillna(0.0)

    long_df = df2.melt(
        id_vars=["customer_raw"],
        value_vars=month_cols,
        var_name="month_name",
        value_name="revenue",
    )

    long_df["month"] = long_df["month_name"].astype(str).str.strip().str.lower().map(MONTH_MAP)

    year = int(sheet_name)
    long_df["year"] = year
    long_df["date"] = pd.to_datetime(dict(year=long_df["year"], month=long_df["month"], day=1))

    long_df["customer"] = long_df["customer_raw"].apply(normalize_customer_name)

    return long_df[["date", "year", "month", "customer_raw", "customer", "revenue"]]


def load_monthly_customer_revenue(xlsx_path: Path) -> pd.DataFrame:
    """Load all year-like sheets from the input workbook into a single long dataframe."""
    xls = pd.ExcelFile(xlsx_path)
    year_sheets = sorted([s for s in xls.sheet_names if re.fullmatch(r"\d{4}", str(s) or "")])

    if not year_sheets:
        raise ValueError("No year sheets found (expected sheet names like 2020, 2021, ...).")

    frames: List[pd.DataFrame] = []
    for s in year_sheets:
        frames.append(parse_year_sheet(xlsx_path, s))

    df_long = pd.concat(frames, ignore_index=True)

    # Trim trailing months where total revenue is 0 (common if the sheet includes future months)
    monthly_totals = df_long.groupby("date")["revenue"].sum().sort_index()
    nonzero = monthly_totals[monthly_totals.abs() > 1e-9]
    if not nonzero.empty:
        last_nonzero = nonzero.index.max()
        df_long = df_long[df_long["date"] <= last_nonzero].copy()

    return df_long


def _rolling_matrix(monthly_matrix: pd.DataFrame, window_months: int) -> pd.DataFrame:
    """
    Rolling trailing sum by customer.

    Uses min_periods=window_months so early periods are NaN (not "partial windows").
    """
    return monthly_matrix.rolling(window=window_months, min_periods=window_months).sum()


def _topk_shares(values: np.ndarray, total: float, ks: Sequence[int]) -> Dict[int, float]:
    """
    Compute top-k share(s) for a 1D array of customer values.
    Returns {k: share}.
    """
    out: Dict[int, float] = {k: np.nan for k in ks}
    if total <= 0 or not np.isfinite(total):
        return out

    # Replace NaN with 0 and clamp very small negatives due to numerical quirks
    v = np.nan_to_num(values, nan=0.0)
    # Sort descending
    v_sorted = np.sort(v)[::-1]
    cumsum = np.cumsum(v_sorted)
    for k in ks:
        if k <= 0:
            out[k] = np.nan
        else:
            kk = min(k, len(v_sorted))
            out[k] = float(cumsum[kk - 1] / total)
    return out


def build_rolling_concentration(
    monthly_matrix: pd.DataFrame,
    windows: Sequence[int] = (12, 24, 36),
    top_ks: Sequence[int] = (1, 5, 10),
) -> pd.DataFrame:
    """Build a time series table of rolling concentration stats."""
    dates = monthly_matrix.index
    total_monthly = monthly_matrix.sum(axis=1)

    out = pd.DataFrame({"date": dates, "total_monthly_revenue": total_monthly.values})

    for w in windows:
        roll = _rolling_matrix(monthly_matrix, w)
        totals = roll.sum(axis=1)
        top1_customer = []
        top1_revenue = []
        top_shares = {k: [] for k in top_ks}

        for d in dates:
            row = roll.loc[d].to_numpy()
            tot = totals.loc[d]
            shares = _topk_shares(row, float(tot) if pd.notna(tot) else float("nan"), top_ks)
            for k, v in shares.items():
                top_shares[k].append(v)

            if pd.isna(tot) or tot == 0:
                top1_customer.append(np.nan)
                top1_revenue.append(np.nan)
            else:
                s = roll.loc[d]
                idx = s.idxmax()
                top1_customer.append(idx)
                top1_revenue.append(float(s.loc[idx]))

        out[f"total_trailing_{w}m"] = totals.values
        out[f"top1_customer_trailing_{w}m"] = top1_customer
        out[f"top1_revenue_trailing_{w}m"] = top1_revenue
        for k in top_ks:
            out[f"top{k}_share_trailing_{w}m"] = top_shares[k]

    return out


def _compute_gap_stats(active_dates: List[pd.Timestamp]) -> Tuple[float, float, int]:
    """
    Given sorted active dates (month starts), return:
      (avg_gap_months, max_gap_months, reactivations)

    - A "gap" is the number of inactive months between two active months.
    - Reactivations = number of times the customer returns after >=1-month gap,
      i.e., (# of active spells - 1)
    """
    if len(active_dates) <= 1:
        return (0.0, 0.0, 0)

    gaps: List[int] = []
    reactivations = 0
    for prev, cur in zip(active_dates[:-1], active_dates[1:]):
        diff_m = (cur.year - prev.year) * 12 + (cur.month - prev.month)
        gap = max(diff_m - 1, 0)
        gaps.append(gap)
        if gap >= 1:
            reactivations += 1

    return (float(np.mean(gaps)), float(np.max(gaps)), int(reactivations))


def build_customer_summary(df_long: pd.DataFrame) -> pd.DataFrame:
    """Build a customer-level summary table for repeat and lumpy behavior."""
    # Ensure consistent sorting
    df_long = df_long.sort_values(["customer", "date"]).copy()

    # Monthly matrix
    monthly_matrix = (
        df_long.pivot_table(index="date", columns="customer", values="revenue", aggfunc="sum", fill_value=0.0)
        .sort_index()
    )

    # Basic revenue aggregates
    lifetime_rev = monthly_matrix.sum(axis=0)

    end_date = monthly_matrix.index.max()
    ttm = _rolling_matrix(monthly_matrix, 12)
    t24 = _rolling_matrix(monthly_matrix, 24)
    t36 = _rolling_matrix(monthly_matrix, 36)

    ttm_last = ttm.loc[end_date] if end_date in ttm.index else pd.Series(dtype=float)
    t24_last = t24.loc[end_date] if end_date in t24.index else pd.Series(dtype=float)
    t36_last = t36.loc[end_date] if end_date in t36.index else pd.Series(dtype=float)

    # Purchase timing
    nonzero = df_long[df_long["revenue"].abs() > 1e-9].copy()
    first_purchase = nonzero.groupby("customer")["date"].min()
    last_purchase = nonzero.groupby("customer")["date"].max()
    active_months = nonzero.groupby("customer")["date"].nunique()
    active_years = nonzero.groupby("customer")["year"].nunique()

    # Quarter activity
    nonzero["quarter"] = nonzero["date"].dt.to_period("Q").astype(str)
    active_quarters = nonzero.groupby("customer")["quarter"].nunique()

    # Tenure months (inclusive, based on first->last purchase)
    tenure_months = (
        (last_purchase.dt.year - first_purchase.dt.year) * 12
        + (last_purchase.dt.month - first_purchase.dt.month)
        + 1
    )

    activity_ratio = active_months / tenure_months

    # Gap / reactivation stats
    gap_avg = {}
    gap_max = {}
    reactivations = {}

    for cust, grp in nonzero.groupby("customer"):
        dates = sorted(grp["date"].unique())
        dates = [pd.Timestamp(d) for d in dates]
        avg_gap, max_gap, react = _compute_gap_stats(dates)
        gap_avg[cust] = avg_gap
        gap_max[cust] = max_gap
        reactivations[cust] = react

    gap_avg_s = pd.Series(gap_avg)
    gap_max_s = pd.Series(gap_max)
    reactivations_s = pd.Series(reactivations)

    # Repeat definition: active in >=2 distinct years
    is_repeat = (active_years >= 2)

    # Peak TTM share per customer
    # Only months where the 12M rolling window is fully defined
    valid_ttm_dates = ttm.sum(axis=1).dropna().index
    ttm_valid = ttm.loc[valid_ttm_dates].copy()
    ttm_totals = ttm_valid.sum(axis=1)
    share = ttm_valid.div(ttm_totals, axis=0)
    peak_ttm_share = share.max(axis=0)

    # Top-10 persistence: fraction of valid TTM months where customer is in the Top 10
    top10_counts = pd.Series(0, index=ttm_valid.columns, dtype=int)
    for d in valid_ttm_dates:
        top10 = ttm_valid.loc[d].nlargest(10).index
        top10_counts.loc[top10] += 1
    top10_persistence = top10_counts / max(len(valid_ttm_dates), 1)

    # Reindex all metrics to match lifetime_rev index
    is_repeat_reindexed = is_repeat.reindex(lifetime_rev.index)
    # Use where to avoid fillna downcasting warning
    is_repeat_reindexed = is_repeat_reindexed.where(is_repeat_reindexed.notna(), False)

    summary = pd.DataFrame({
        "customer": lifetime_rev.index,
        "lifetime_revenue": lifetime_rev.values,
        "ttm_revenue_last": ttm_last.reindex(lifetime_rev.index).values,
        "t24m_revenue_last": t24_last.reindex(lifetime_rev.index).values,
        "t36m_revenue_last": t36_last.reindex(lifetime_rev.index).values,
        "first_purchase_month": first_purchase.reindex(lifetime_rev.index).values,
        "last_purchase_month": last_purchase.reindex(lifetime_rev.index).values,
        "active_months": active_months.reindex(lifetime_rev.index).fillna(0).astype(int).values,
        "active_quarters": active_quarters.reindex(lifetime_rev.index).fillna(0).astype(int).values,
        "active_years": active_years.reindex(lifetime_rev.index).fillna(0).astype(int).values,
        "tenure_months": tenure_months.reindex(lifetime_rev.index).fillna(0).astype(int).values,
        "activity_ratio": activity_ratio.reindex(lifetime_rev.index).fillna(0.0).values,
        "avg_gap_months": gap_avg_s.reindex(lifetime_rev.index).fillna(0.0).values,
        "max_gap_months": gap_max_s.reindex(lifetime_rev.index).fillna(0.0).values,
        "reactivations": reactivations_s.reindex(lifetime_rev.index).fillna(0).astype(int).values,
        "is_repeat_customer": is_repeat_reindexed.values,
        "peak_ttm_share": peak_ttm_share.reindex(lifetime_rev.index).values,
        "top10_ttm_persistence": top10_persistence.reindex(lifetime_rev.index).values,
    })

    summary = summary.sort_values("lifetime_revenue", ascending=False).reset_index(drop=True)
    return summary


def build_customer_master(df_long: pd.DataFrame) -> pd.DataFrame:
    """
    Build a customer master sheet identifying potential duplicate names.

    Returns a dataframe with all customer variations and suggested consolidations.
    """
    # Get all unique customer names (normalized and raw)
    customer_mapping = (
        df_long[["customer", "customer_raw"]]
        .drop_duplicates()
        .groupby("customer")["customer_raw"]
        .apply(list)
        .to_dict()
    )

    # Calculate revenue per normalized customer
    revenue_by_customer = (
        df_long.groupby("customer")["revenue"].sum().to_dict()
    )

    # Find potential duplicates
    all_customers = list(customer_mapping.keys())
    duplicate_matches = find_potential_duplicates(all_customers)

    # Build groups of related customers
    # Only group HIGH confidence matches to avoid false positives from transitive grouping
    customer_to_group: Dict[str, int] = {}
    group_counter = 0

    for match in duplicate_matches:
        # Only create groups for HIGH confidence matches to avoid false positives
        if match.confidence != "HIGH":
            continue

        cust1_group = customer_to_group.get(match.customer1)
        cust2_group = customer_to_group.get(match.customer2)

        if cust1_group is None and cust2_group is None:
            # Neither in a group yet - create new group
            customer_to_group[match.customer1] = group_counter
            customer_to_group[match.customer2] = group_counter
            group_counter += 1
        elif cust1_group is not None and cust2_group is None:
            # Add cust2 to cust1's group
            customer_to_group[match.customer2] = cust1_group
        elif cust2_group is not None and cust1_group is None:
            # Add cust1 to cust2's group
            customer_to_group[match.customer1] = cust2_group
        elif cust1_group != cust2_group:
            # Merge groups (assign all members of cust2's group to cust1's group)
            for cust, grp in list(customer_to_group.items()):
                if grp == cust2_group:
                    customer_to_group[cust] = cust1_group

    # Build the master dataframe
    # Include all customers and show ALL potential matches (grouped and ungrouped)
    rows = []

    # Create a mapping of customer -> all their matches
    customer_matches: Dict[str, List[DuplicateMatch]] = {}
    for match in duplicate_matches:
        if match.customer1 not in customer_matches:
            customer_matches[match.customer1] = []
        if match.customer2 not in customer_matches:
            customer_matches[match.customer2] = []
        customer_matches[match.customer1].append(match)
        customer_matches[match.customer2].append(match)

    for normalized_customer, raw_variations in customer_mapping.items():
        group_id = customer_to_group.get(normalized_customer, -1)
        revenue = revenue_by_customer.get(normalized_customer, 0.0)

        # Get all matches for this customer
        matches = customer_matches.get(normalized_customer, [])

        # Calculate combined revenue for the group (HIGH confidence only)
        combined_revenue = revenue
        if group_id >= 0:
            for other_cust, other_group in customer_to_group.items():
                if other_group == group_id and other_cust != normalized_customer:
                    combined_revenue += revenue_by_customer.get(other_cust, 0.0)

        # Determine suggested master name
        suggested_master = normalized_customer
        if group_id >= 0:
            # Get all customers in this group
            group_members = [c for c, g in customer_to_group.items() if g == group_id]
            # Pick the one with highest revenue as the master
            suggested_master = max(group_members, key=lambda c: revenue_by_customer.get(c, 0.0))

        # Add rows for each match (or one row if no matches)
        if matches:
            for match in matches:
                # Get the other customer in this match
                other_customer = match.customer2 if match.customer1 == normalized_customer else match.customer1

                # Add row for each raw variation
                for raw_name in raw_variations:
                    rows.append({
                        "duplicate_group_id": group_id if group_id >= 0 else None,
                        "customer_normalized": normalized_customer,
                        "customer_raw": raw_name,
                        "potential_duplicate": other_customer,
                        "suggested_master_name": suggested_master if group_id >= 0 else None,
                        "confidence": match.confidence,
                        "merge_reason": match.reason,
                        "similarity_score": match.similarity,
                    })
        else:
            # No matches - just add customer info
            for raw_name in raw_variations:
                rows.append({
                    "duplicate_group_id": None,
                    "customer_normalized": normalized_customer,
                    "customer_raw": raw_name,
                    "potential_duplicate": None,
                    "suggested_master_name": None,
                    "confidence": None,
                    "merge_reason": None,
                    "similarity_score": None,
                })

    master_df = pd.DataFrame(rows)

    # Sort: duplicates first (by group_id), then alphabetically
    # Split into duplicates and non-duplicates for separate sorting
    duplicates = master_df[master_df["duplicate_group_id"].notna()].copy()
    non_duplicates = master_df[master_df["duplicate_group_id"].isna()].copy()

    # Sort duplicates by group_id
    if len(duplicates) > 0:
        duplicates = duplicates.sort_values("duplicate_group_id")

    # Sort non-duplicates alphabetically by customer_normalized
    if len(non_duplicates) > 0:
        non_duplicates = non_duplicates.sort_values("customer_normalized")

    # Combine: duplicates first, then alphabetical
    master_df = pd.concat([duplicates, non_duplicates], ignore_index=True)

    return master_df


def _format_excel(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame, freeze_panes: Tuple[int, int] = (1, 0)) -> None:
    """Apply lightweight formatting via openpyxl (freeze header row, autofilter, number formats)."""
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import numbers

    ws = writer.sheets[sheet_name]
    ws.freeze_panes = ws.cell(row=freeze_panes[0] + 1, column=freeze_panes[1] + 1)  # type: ignore[attr-defined]
    ws.auto_filter.ref = ws.dimensions  # type: ignore[attr-defined]

    # Basic column width auto-fit (capped); only scan the first ~200 rows for speed
    for col_idx, col in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        col_name = str(col).lower()

        try:
            max_len = max([len(str(col))] + [len(str(x)) for x in df[col].head(200).tolist()])
        except Exception:
            max_len = len(str(col))
        width = min(max(10, max_len + 2), 45)
        ws.column_dimensions[col_letter].width = width  # type: ignore[attr-defined]

        # Apply number formatting based on column name
        num_format = None

        # Percentage/share/ratio columns: Percentage with 0 decimals (check first to avoid conflicts)
        if 'share' in col_name or 'ratio' in col_name or 'persistence' in col_name:
            num_format = '0%'

        # Date columns: YYYY-MM-DD format (specific date column names only)
        elif col_name == 'date' or col_name.endswith('_month') or col_name.endswith('_date'):
            num_format = 'yyyy-mm-dd'

        # Gap columns: Number with 1 decimal place
        elif 'gap' in col_name:
            num_format = '0.0'

        # Revenue columns: Accounting format with 0 decimals
        elif 'revenue' in col_name or 'total_trailing' in col_name or col_name.startswith('total_'):
            num_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'

        # Apply the format to all data cells in this column (skip header row)
        if num_format:
            for row in range(2, ws.max_row + 1):
                ws[f'{col_letter}{row}'].number_format = num_format  # type: ignore[attr-defined]


def _add_customer_segmentation_columns(writer: pd.ExcelWriter, df_long: pd.DataFrame) -> None:
    """Add formula-based segmentation columns to Customer_Summary sheet."""
    from openpyxl.utils import get_column_letter

    ws = writer.sheets["Customer_Summary"]

    # Get the most recent month from the data
    most_recent_month = pd.to_datetime(df_long['date']).max()
    most_recent_date_str = most_recent_month.strftime('%Y-%m-%d')

    # Find column indices (1-based) for existing columns
    # Expected columns: customer, lifetime_revenue, ttm_revenue_last, t24m_revenue_last,
    # t36m_revenue_last, first_purchase_month, last_purchase_month, active_months, active_quarters,
    # active_years, tenure_months, activity_ratio, avg_gap_months, max_gap_months, reactivations,
    # is_repeat_customer, peak_ttm_share, top10_ttm_persistence

    # Column positions (A=1, B=2, etc.)
    col_last_purchase = 7  # last_purchase_month is column G
    col_ttm = 3  # ttm_revenue_last is column C
    col_lifetime = 2  # lifetime_revenue is column B
    col_tenure = 11  # tenure_months is column K
    col_peak_share = 17  # peak_ttm_share is column Q

    # Add new column headers
    num_rows = ws.max_row
    next_col = ws.max_column + 1

    headers = [
        "months_since_last_purchase",
        "status",
        "is_high_value",
        "segment"
    ]

    for i, header in enumerate(headers):
        col_letter = get_column_letter(next_col + i)
        ws[f"{col_letter}1"] = header
        ws.column_dimensions[col_letter].width = 20  # type: ignore[attr-defined]

    # Add formulas for each data row (starting at row 2)
    for row in range(2, num_rows + 1):
        col_idx = next_col

        # Column 1: months_since_last_purchase
        # =DATEDIF(H2, DATE(2025,9,1), "M")
        last_purchase_cell = f"{get_column_letter(col_last_purchase)}{row}"
        formula1 = f'=DATEDIF({last_purchase_cell},DATE({most_recent_month.year},{most_recent_month.month},1),"M")'
        ws[f"{get_column_letter(col_idx)}{row}"] = formula1

        col_idx += 1

        # Column 2: status (Active/Inactive/Churned)
        # =IF(T2<=6,"Active",IF(T2<=18,"Inactive","Churned"))
        months_since_col = get_column_letter(col_idx - 1)
        formula2 = f'=IF({months_since_col}{row}<=6,"Active",IF({months_since_col}{row}<=18,"Inactive","Churned"))'
        ws[f"{get_column_letter(col_idx)}{row}"] = formula2

        col_idx += 1

        # Column 3: is_high_value
        # =OR(C2>=1000000,R2>=0.02)
        lifetime_col = get_column_letter(col_lifetime)
        peak_share_col = get_column_letter(col_peak_share)
        formula3 = f'=OR({lifetime_col}{row}>=1000000,{peak_share_col}{row}>=0.02)'
        ws[f"{get_column_letter(col_idx)}{row}"] = formula3

        col_idx += 1

        # Column 4: segment (for active customers only)
        # =IF(U2="Active",IF(AND(D2>=2000000,OR(L2>=36,C2>=5000000)),"Strategic",IF(D2>=1000000,"High Value","Mid Value")),"")
        status_col = get_column_letter(next_col + 1)  # status column
        ttm_col = get_column_letter(col_ttm)
        tenure_col = get_column_letter(col_tenure)
        lifetime_col = get_column_letter(col_lifetime)
        formula4 = f'=IF({status_col}{row}="Active",IF(AND({ttm_col}{row}>=2000000,OR({tenure_col}{row}>=36,{lifetime_col}{row}>=5000000)),"Strategic",IF({ttm_col}{row}>=1000000,"High Value","Mid Value")),"")'
        ws[f"{get_column_letter(col_idx)}{row}"] = formula4

    # Update autofilter to include new columns
    ws.auto_filter.ref = ws.dimensions  # type: ignore[attr-defined]



def generate_customer_master_file(df_long: pd.DataFrame, output_path: Path) -> pd.DataFrame:
    """
    Generate a customer master Excel file template for manual editing.

    Returns a dataframe with:
    - customer_normalized: The auto-normalized customer name
    - customer_master: Name to use in analysis (user edits this to consolidate)
    - suggested_consolidation: Suggested master name from duplicate detection
    - notes: Comments about potential duplicates

    The Excel file is written next to the output file with _master.xlsx suffix.
    """
    # Get all unique customers and their suggested consolidations
    customer_mapping = (
        df_long[["customer", "customer_raw"]]
        .drop_duplicates()
        .groupby("customer")["customer_raw"]
        .agg(lambda x: ", ".join(sorted(set(x))))
        .to_dict()
    )

    # Get duplicate information
    all_customers = list(customer_mapping.keys())
    duplicate_matches = find_potential_duplicates(all_customers)

    # Build a mapping of customer -> suggested master (from HIGH confidence groups)
    customer_to_group: Dict[str, int] = {}
    group_counter = 0

    for match in duplicate_matches:
        if match.confidence != "HIGH":
            continue

        cust1_group = customer_to_group.get(match.customer1)
        cust2_group = customer_to_group.get(match.customer2)

        if cust1_group is None and cust2_group is None:
            customer_to_group[match.customer1] = group_counter
            customer_to_group[match.customer2] = group_counter
            group_counter += 1
        elif cust1_group is not None and cust2_group is None:
            customer_to_group[match.customer2] = cust1_group
        elif cust2_group is not None and cust1_group is None:
            customer_to_group[match.customer1] = cust2_group
        elif cust1_group != cust2_group:
            for cust, grp in list(customer_to_group.items()):
                if grp == cust2_group:
                    customer_to_group[cust] = cust1_group

    # Calculate revenue per customer to pick the master name
    revenue_by_customer = df_long.groupby("customer")["revenue"].sum().to_dict()

    # Determine suggested master for each group
    group_masters: Dict[int, str] = {}
    for cust, group_id in customer_to_group.items():
        if group_id not in group_masters:
            group_members = [c for c, g in customer_to_group.items() if g == group_id]
            group_masters[group_id] = max(group_members, key=lambda c: revenue_by_customer.get(c, 0.0))

    # Create a mapping of customer -> all their matches
    customer_matches: Dict[str, List[DuplicateMatch]] = {}
    for match in duplicate_matches:
        if match.customer1 not in customer_matches:
            customer_matches[match.customer1] = []
        if match.customer2 not in customer_matches:
            customer_matches[match.customer2] = []
        customer_matches[match.customer1].append(match)
        customer_matches[match.customer2].append(match)

    # Build master rows - one row per potential match (or one row if no matches)
    rows = []
    for customer in sorted(all_customers):
        group_id = customer_to_group.get(customer)
        suggested_master = group_masters.get(group_id) if group_id is not None else customer

        # Get all matches for this customer
        matches = customer_matches.get(customer, [])

        if matches:
            # Create a row for each potential duplicate
            for match in matches:
                other_customer = match.customer2 if match.customer1 == customer else match.customer1
                rows.append({
                    "customer_normalized": customer,
                    "customer_master": customer,  # User edits this column
                    "suggested_consolidation": suggested_master if group_id is not None else "",
                    "raw_variations": customer_mapping[customer],
                    "potential_duplicate": other_customer,
                    "confidence": match.confidence,
                    "merge_reason": match.reason,
                    "similarity_score": match.similarity,
                })
        else:
            # No matches - single row
            rows.append({
                "customer_normalized": customer,
                "customer_master": customer,  # User edits this column
                "suggested_consolidation": "",
                "raw_variations": customer_mapping[customer],
                "potential_duplicate": None,
                "confidence": None,
                "merge_reason": None,
                "similarity_score": None,
            })

    master_df = pd.DataFrame(rows)

    # Sort: customers with potential duplicates first, then no duplicates
    # Add visual separation between the two groups
    with_duplicates = master_df[master_df["potential_duplicate"].notna()].copy()
    without_duplicates = master_df[master_df["potential_duplicate"].isna()].copy()

    # Add 3 empty rows as visual separator with matching dtypes
    separator_rows = pd.DataFrame(
        [[None] * len(master_df.columns)] * 3,
        columns=master_df.columns
    ).astype(master_df.dtypes.to_dict())

    # Combine: duplicates, separator, non-duplicates
    master_df = pd.concat([with_duplicates, separator_rows, without_duplicates], ignore_index=True)

    # Write Excel file with formatting
    xlsx_path = output_path.parent / "customer_master.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        master_df.to_excel(writer, sheet_name="Customer_Master", index=False)
        _format_excel(writer, "Customer_Master", master_df)

        # Apply percentage format to similarity_score column
        from openpyxl.styles import numbers
        ws = writer.sheets["Customer_Master"]
        similarity_col_idx = list(master_df.columns).index("similarity_score") + 1
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(similarity_col_idx)

        # Format as percentage with 0 decimal places (skip header row)
        for row in range(2, len(master_df) + 2):
            cell = ws[f"{col_letter}{row}"]
            cell.number_format = numbers.FORMAT_PERCENTAGE

    print(f"Generated customer master template: {xlsx_path}")

    return master_df


def apply_customer_master(df_long: pd.DataFrame, master_path: Path) -> pd.DataFrame:
    """
    Apply customer master mapping to consolidate duplicate customer names.

    Reads the master Excel file and replaces customer names according to the mapping.
    """
    if not master_path.exists():
        raise FileNotFoundError(f"Customer master file not found: {master_path}")

    master_df = pd.read_excel(master_path, engine="openpyxl")

    # Validate required columns
    required_cols = ["customer_normalized", "customer_master"]
    missing = set(required_cols) - set(master_df.columns)
    if missing:
        raise ValueError(f"Customer master file missing required columns: {missing}")

    # Build mapping dictionary
    mapping = dict(zip(master_df["customer_normalized"], master_df["customer_master"]))

    # Apply mapping to df_long
    df_long = df_long.copy()
    df_long["customer"] = df_long["customer"].map(lambda c: mapping.get(c, c))

    print(f"Applied customer master mapping from: {master_path}")
    return df_long


def write_analysis_workbook(df_long: pd.DataFrame, output_path: Path) -> None:
    """Generate the analysis workbook."""
    # Build core artifacts
    monthly_matrix = (
        df_long.pivot_table(index="date", columns="customer", values="revenue", aggfunc="sum", fill_value=0.0)
        .sort_index()
    )

    rolling_conc = build_rolling_concentration(monthly_matrix, windows=(12, 24, 36), top_ks=(1, 5, 10))

    cust_summary = build_customer_summary(df_long)

    # Top tables
    top25_lifetime = cust_summary.sort_values("lifetime_revenue", ascending=False).head(25)
    top25_ttm = cust_summary.sort_values("ttm_revenue_last", ascending=False).head(25)
    top25_peak = cust_summary.sort_values("peak_ttm_share", ascending=False).head(25)

    # Write workbook
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Customer summary first
        cust_summary.to_excel(writer, sheet_name="Customer_Summary", index=False)
        _format_excel(writer, "Customer_Summary", cust_summary)
        # Add segmentation columns with formulas
        _add_customer_segmentation_columns(writer, df_long)

        # Keep the long data so future users can extend analysis
        # Exclude customer_raw to keep only the normalized/master customer name
        df_long_output = df_long.sort_values(["date", "customer"]).drop(columns=["customer_raw"], errors="ignore")
        df_long_output.to_excel(writer, sheet_name="Monthly_Long", index=False)
        _format_excel(writer, "Monthly_Long", df_long_output)

        monthly_matrix.to_excel(writer, sheet_name="Monthly_Matrix")
        _format_excel(writer, "Monthly_Matrix", monthly_matrix.reset_index())

        rolling_conc.to_excel(writer, sheet_name="Rolling_Concentration", index=False)
        _format_excel(writer, "Rolling_Concentration", rolling_conc)

        top25_lifetime.to_excel(writer, sheet_name="Top25_Lifetime", index=False)
        _format_excel(writer, "Top25_Lifetime", top25_lifetime)

        top25_ttm.to_excel(writer, sheet_name="Top25_TTM", index=False)
        _format_excel(writer, "Top25_TTM", top25_ttm)

        top25_peak.to_excel(writer, sheet_name="Top25_PeakTTMShare", index=False)
        _format_excel(writer, "Top25_PeakTTMShare", top25_peak)


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Build repeat-customer purchase-pattern analysis workbook.")
    parser.add_argument("--input", required=True, help="Path to the source Excel file (month-wise customer revenue).")
    parser.add_argument("--output", required=True, help="Path to write the analysis Excel workbook.")
    parser.add_argument(
        "--master",
        help="Path to customer master Excel file (.xlsx) for consolidating duplicate names. "
        "If not provided, a template will be generated alongside the output file."
    )
    args = parser.parse_args(argv)

    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()
    master_path = Path(args.master).expanduser().resolve() if args.master else None

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    # Load raw data
    df_long = load_monthly_customer_revenue(input_path)

    # Apply customer master mapping if provided
    if master_path:
        df_long = apply_customer_master(df_long, master_path)
    else:
        # Only generate customer master template on first run (without --master)
        # This prevents overwriting user's edited mappings
        generate_customer_master_file(df_long, output_path)

    # Generate analysis workbook
    write_analysis_workbook(df_long, output_path)

    print(f"Wrote analysis workbook: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
