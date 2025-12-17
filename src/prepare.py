#!/usr/bin/env python3
"""
Prepare customer revenue data for analysis.

Reads a multi-sheet Excel workbook with year tabs containing "Name of Customer" rows
and month columns, normalizes customer names, detects duplicates, and outputs:
1. A prepared revenue file with flattened, normalized data
2. A customer master file for managing duplicate consolidation

Usage:
  python prepare.py --input "4.0.5 Financial - Month-wise Customer Revenue.xlsx"
  python prepare.py --input "4.0.5 Financial..." --master "4.0.4 Financial...customers.xlsx"
"""

from __future__ import annotations

import argparse
import math
import re
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd


MONTH_MAP: Dict[str, int] = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}


# -----------------------------------------------------------------------------
# Customer Name Normalization
# -----------------------------------------------------------------------------

def normalize_customer_name(name: str) -> str:
    """Normalize customer names for aggregation (case + whitespace)."""
    s = str(name).strip()
    s = re.sub(r"\s+", " ", s)
    return s.upper()


def _get_distinctive_name(name: str) -> str:
    """
    Extract the distinctive part of a company name (before generic suffixes).

    For "ALIGNED TECHNOLOGIES INC." returns "ALIGNED"
    For "SUN LIFE ASSURANCE COMPANY" returns "SUN LIFE ASSURANCE"
    """
    generic_terms = [
        'inc', 'incorporated', 'corp', 'corporation', 'ltd', 'limited',
        'llc', 'lp', 'ulc', 'plc', 'llp', 'company', 'co',
        'technologies', 'technology', 'services', 'group', 'partners',
        'canada', 'canadian', 'international', 'global',
    ]

    tokens = name.lower().split()
    distinctive = []
    for i, token in enumerate(tokens):
        clean_token = token.strip('.,')
        if i == 0 or clean_token not in generic_terms:
            distinctive.append(token)

    return ' '.join(distinctive) if distinctive else name.lower()


def _similarity_ratio(a: str, b: str) -> float:
    """Calculate similarity ratio between two strings (0.0 to 1.0)."""
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


def _strip_legal_suffix(name: str) -> str:
    """Remove common legal suffixes for comparison."""
    suffixes = [
        r'\s+inc\.?$', r'\s+incorporated$', r'\s+corp\.?$', r'\s+corporation$',
        r'\s+ltd\.?$', r'\s+limited$', r'\s+llc$', r'\s+lp$', r'\s+ulc$',
        r'\s+plc$', r'\s+llp\.?$',
    ]

    result = name.lower()
    for suffix_pattern in suffixes:
        result = re.sub(suffix_pattern, '', result, flags=re.IGNORECASE)

    return result.strip()


# -----------------------------------------------------------------------------
# Duplicate Detection
# -----------------------------------------------------------------------------

@dataclass
class DuplicateMatch:
    """Represents a potential duplicate customer relationship."""
    customer1: str
    customer2: str
    similarity: float
    confidence: str  # HIGH, MEDIUM, LOW
    reason: str


def find_potential_duplicates(customers: List[str]) -> List[DuplicateMatch]:
    """
    Find potential duplicate customer names using multiple heuristics.

    Args:
        customers: List of normalized customer names

    Returns:
        List of DuplicateMatch objects for potential duplicates
    """
    matches: List[DuplicateMatch] = []
    customers_sorted = sorted(set(customers))

    for i, cust1 in enumerate(customers_sorted):
        for cust2 in customers_sorted[i + 1:]:
            sim = _similarity_ratio(cust1, cust2)
            dist1 = _get_distinctive_name(cust1)
            dist2 = _get_distinctive_name(cust2)
            dist_sim = _similarity_ratio(dist1, dist2)

            confidence = None
            reason = None

            # Pattern 1: One name contains the other
            if cust1.lower() in cust2.lower() or cust2.lower() in cust1.lower():
                if sim > 0.85:
                    confidence = "HIGH"
                    reason = "One name contains the other"

            # Pattern 2: Only differ by legal suffix
            if confidence is None:
                stripped1 = _strip_legal_suffix(cust1)
                stripped2 = _strip_legal_suffix(cust2)
                if stripped1 == stripped2 and stripped1:
                    confidence = "HIGH"
                    reason = "Legal suffix variation only"

            # Pattern 3: Very high similarity
            if confidence is None and sim > 0.90 and dist_sim > 0.60:
                confidence = "HIGH"
                reason = f"Very high similarity ({sim:.2f})"

            # Pattern 4: High similarity with common patterns
            if confidence is None and sim > 0.75 and dist_sim > 0.50:
                location_pattern = r'\s+-\s+\d+\s+.*(?:st|street|ave|avenue|rd|road)\.?$'
                if re.search(location_pattern, cust1.lower()) or re.search(location_pattern, cust2.lower()):
                    confidence = "MEDIUM"
                    reason = "Location suffix variation"
                elif "(" in cust1 or "(" in cust2:
                    confidence = "MEDIUM"
                    reason = "Parenthetical addition"
                else:
                    confidence = "MEDIUM"
                    reason = f"High similarity ({sim:.2f})"

            # Pattern 5: Moderate similarity
            if confidence is None and sim > 0.70 and dist_sim > 0.50:
                confidence = "LOW"
                reason = f"Moderate similarity ({sim:.2f})"

            if confidence:
                matches.append(DuplicateMatch(
                    customer1=cust1,
                    customer2=cust2,
                    similarity=sim,
                    confidence=confidence,
                    reason=reason
                ))

    return matches


# -----------------------------------------------------------------------------
# Input Parsing
# -----------------------------------------------------------------------------

def _find_header_row(df_raw: pd.DataFrame, max_scan_rows: int = 60) -> Optional[int]:
    """Find the row index containing 'Name of Customer'."""
    scan_n = min(len(df_raw), max_scan_rows)
    for i in range(scan_n):
        row = df_raw.iloc[i].astype(str).str.strip().str.lower()
        if (row == "name of customer").any():
            return i
    return None


def _month_cols_from_headers(headers: Sequence[object]) -> List[str]:
    """Return the month columns present in the header row."""
    month_cols: List[str] = []
    for h in headers:
        key = str(h).strip().lower()
        if key in MONTH_MAP:
            month_cols.append(str(h))
    return month_cols


def parse_year_sheet(xlsx_path: Path, sheet_name: str) -> pd.DataFrame:
    """
    Parse a year sheet into long format:
      date, year, month, customer_raw, customer, revenue
    """
    df_raw = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=None)
    header_row = _find_header_row(df_raw)
    if header_row is None:
        raise ValueError(f"Could not find 'Name of Customer' header in sheet '{sheet_name}'.")

    headers = df_raw.iloc[header_row].tolist()
    headers = [
        (h if not (isinstance(h, float) and math.isnan(h)) else f"unnamed_{j}")
        for j, h in enumerate(headers)
    ]

    df = df_raw.iloc[header_row + 1:].copy()
    df.columns = headers
    df = df.dropna(how="all")

    cust_col = next((c for c in df.columns if str(c).strip().lower() == "name of customer"), df.columns[0])

    month_cols: List[str] = []
    for c in df.columns:
        key = str(c).strip().lower()
        if key in MONTH_MAP:
            month_cols.append(c)

    if not month_cols:
        raise ValueError(f"No month columns recognized in sheet '{sheet_name}'.")

    df2 = df[[cust_col] + month_cols].copy()
    df2 = df2.rename(columns={cust_col: "customer_raw"})

    df2["customer_raw"] = df2["customer_raw"].astype(str).str.strip()
    df2.loc[df2["customer_raw"].str.lower().isin(["nan", ""]), "customer_raw"] = np.nan
    df2 = df2.dropna(subset=["customer_raw"])
    df2 = df2[~df2["customer_raw"].str.lower().isin(["total", "totals", "grand total"])]

    for c in month_cols:
        col_series = df2[c].astype(str)
        col_series = col_series.replace(["-", "–", "—", ""], np.nan)
        df2[c] = pd.to_numeric(col_series, errors="coerce").fillna(0.0)

    long_df = df2.melt(
        id_vars=["customer_raw"],
        value_vars=month_cols,
        var_name="month_name",
        value_name="revenue",
    )

    long_df["month"] = long_df["month_name"].astype(str).str.strip().str.lower().map(MONTH_MAP)

    year = int(sheet_name)
    long_df["year"] = year
    long_df["date"] = pd.to_datetime(dict(year=long_df["year"], month=long_df["month"], day=1))
    long_df["customer"] = long_df["customer_raw"].apply(normalize_customer_name)

    return long_df[["date", "year", "month", "customer_raw", "customer", "revenue"]]


def load_monthly_customer_revenue(xlsx_path: Path) -> pd.DataFrame:
    """Load all year-like sheets from the input workbook into a single long dataframe."""
    xls = pd.ExcelFile(xlsx_path)
    year_sheets = sorted([s for s in xls.sheet_names if re.fullmatch(r"\d{4}", str(s) or "")])

    if not year_sheets:
        raise ValueError("No year sheets found (expected sheet names like 2020, 2021, ...).")

    frames: List[pd.DataFrame] = []
    for s in year_sheets:
        frames.append(parse_year_sheet(xlsx_path, s))

    df_long = pd.concat(frames, ignore_index=True)

    # Trim trailing months where total revenue is 0
    monthly_totals = df_long.groupby("date")["revenue"].sum().sort_index()
    nonzero = monthly_totals[monthly_totals.abs() > 1e-9]
    if not nonzero.empty:
        last_nonzero = nonzero.index.max()
        df_long = df_long[df_long["date"] <= last_nonzero].copy()

    return df_long


# -----------------------------------------------------------------------------
# Customer Master Management
# -----------------------------------------------------------------------------

def generate_customer_master(
    df_long: pd.DataFrame,
    existing_master: Optional[pd.DataFrame] = None
) -> pd.DataFrame:
    """
    Generate a customer master dataframe for managing duplicate consolidation.

    If existing_master is provided, preserves those mappings and adds new customers.
    """
    # Get all unique customers
    customer_mapping = (
        df_long[["customer", "customer_raw"]]
        .drop_duplicates()
        .groupby("customer")["customer_raw"]
        .agg(lambda x: ", ".join(sorted(set(x))))
        .to_dict()
    )

    all_customers = list(customer_mapping.keys())
    duplicate_matches = find_potential_duplicates(all_customers)

    # Build groups for HIGH confidence matches
    customer_to_group: Dict[str, int] = {}
    group_counter = 0

    for match in duplicate_matches:
        if match.confidence != "HIGH":
            continue

        cust1_group = customer_to_group.get(match.customer1)
        cust2_group = customer_to_group.get(match.customer2)

        if cust1_group is None and cust2_group is None:
            customer_to_group[match.customer1] = group_counter
            customer_to_group[match.customer2] = group_counter
            group_counter += 1
        elif cust1_group is not None and cust2_group is None:
            customer_to_group[match.customer2] = cust1_group
        elif cust2_group is not None and cust1_group is None:
            customer_to_group[match.customer1] = cust2_group
        elif cust1_group != cust2_group:
            for cust, grp in list(customer_to_group.items()):
                if grp == cust2_group:
                    customer_to_group[cust] = cust1_group

    # Calculate revenue per customer to pick master names
    revenue_by_customer = df_long.groupby("customer")["revenue"].sum().to_dict()

    # Determine suggested master for each group
    group_masters: Dict[int, str] = {}
    for cust, group_id in customer_to_group.items():
        if group_id not in group_masters:
            group_members = [c for c, g in customer_to_group.items() if g == group_id]
            group_masters[group_id] = max(group_members, key=lambda c: revenue_by_customer.get(c, 0.0))

    # Build customer -> matches mapping
    customer_matches: Dict[str, List[DuplicateMatch]] = {}
    for match in duplicate_matches:
        if match.customer1 not in customer_matches:
            customer_matches[match.customer1] = []
        if match.customer2 not in customer_matches:
            customer_matches[match.customer2] = []
        customer_matches[match.customer1].append(match)
        customer_matches[match.customer2].append(match)

    # Load existing mappings if provided
    existing_mappings: Dict[str, str] = {}
    if existing_master is not None:
        for _, row in existing_master.iterrows():
            if pd.notna(row.get("customer_normalized")) and pd.notna(row.get("customer_master")):
                existing_mappings[row["customer_normalized"]] = row["customer_master"]

    # Build master rows
    rows = []
    for customer in sorted(all_customers):
        group_id = customer_to_group.get(customer)
        suggested_master = group_masters.get(group_id) if group_id is not None else customer

        # Use existing mapping if available, otherwise default to customer name
        if customer in existing_mappings:
            master_name = existing_mappings[customer]
            is_new = False
        else:
            master_name = customer
            is_new = True

        matches = customer_matches.get(customer, [])

        if matches:
            for match in matches:
                other_customer = match.customer2 if match.customer1 == customer else match.customer1
                rows.append({
                    "customer_normalized": customer,
                    "customer_master": master_name,
                    "suggested_consolidation": suggested_master if group_id is not None else "",
                    "raw_variations": customer_mapping[customer],
                    "potential_duplicate": other_customer,
                    "confidence": match.confidence,
                    "merge_reason": match.reason,
                    "similarity_score": match.similarity,
                    "is_new": is_new,
                })
        else:
            rows.append({
                "customer_normalized": customer,
                "customer_master": master_name,
                "suggested_consolidation": "",
                "raw_variations": customer_mapping[customer],
                "potential_duplicate": None,
                "confidence": None,
                "merge_reason": None,
                "similarity_score": None,
                "is_new": is_new,
            })

    master_df = pd.DataFrame(rows)

    # Sort: duplicates first, then no duplicates
    with_duplicates = master_df[master_df["potential_duplicate"].notna()].copy()
    without_duplicates = master_df[master_df["potential_duplicate"].isna()].copy()

    if len(with_duplicates) > 0:
        # Sort by is_new (new ones first for review), then by customer_normalized
        with_duplicates = with_duplicates.sort_values(["is_new", "customer_normalized"], ascending=[False, True])

    if len(without_duplicates) > 0:
        without_duplicates = without_duplicates.sort_values(["is_new", "customer_normalized"], ascending=[False, True])

    # Combine with separator (use empty strings to avoid FutureWarning about all-NA concat)
    separator_rows = pd.DataFrame(
        [[""] * len(master_df.columns)] * 3,
        columns=master_df.columns
    )

    master_df = pd.concat([with_duplicates, separator_rows, without_duplicates], ignore_index=True)

    return master_df


def apply_customer_master(df_long: pd.DataFrame, master_df: pd.DataFrame) -> pd.DataFrame:
    """Apply customer master mapping to consolidate duplicate customer names."""
    # Build mapping dictionary (skip empty rows)
    mapping = {}
    for _, row in master_df.iterrows():
        if pd.notna(row.get("customer_normalized")) and pd.notna(row.get("customer_master")):
            mapping[row["customer_normalized"]] = row["customer_master"]

    df_long = df_long.copy()
    df_long["customer"] = df_long["customer"].map(lambda c: mapping.get(c, c))

    return df_long


# -----------------------------------------------------------------------------
# Excel Output
# -----------------------------------------------------------------------------

def _format_excel(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    """Apply formatting to Excel sheet."""
    from openpyxl.utils import get_column_letter

    ws = writer.sheets[sheet_name]
    ws.freeze_panes = ws.cell(row=2, column=1)
    ws.auto_filter.ref = ws.dimensions

    for col_idx, col in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        col_name = str(col).lower()

        try:
            max_len = max([len(str(col))] + [len(str(x)) for x in df[col].head(200).tolist()])
        except Exception:
            max_len = len(str(col))
        width = min(max(10, max_len + 2), 45)
        ws.column_dimensions[col_letter].width = width

        num_format = None
        if 'share' in col_name or 'ratio' in col_name or 'similarity' in col_name:
            num_format = '0%'
        elif col_name == 'date' or col_name.endswith('_month') or col_name.endswith('_date'):
            num_format = 'yyyy-mm-dd'
        elif 'revenue' in col_name:
            num_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'

        if num_format:
            for row in range(2, ws.max_row + 1):
                ws[f'{col_letter}{row}'].number_format = num_format


def write_prepared_file(df_long: pd.DataFrame, output_path: Path) -> None:
    """Write the prepared revenue data to Excel."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Output only the columns needed for analytics
    df_output = df_long[["date", "year", "month", "customer", "revenue"]].copy()
    df_output = df_output.sort_values(["date", "customer"])

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_output.to_excel(writer, sheet_name="Revenue_Detail", index=False)
        _format_excel(writer, "Revenue_Detail", df_output)

    print(f"Wrote prepared data: {output_path}")


def write_customer_master_file(master_df: pd.DataFrame, output_path: Path) -> None:
    """Write the customer master file to Excel."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        master_df.to_excel(writer, sheet_name="Customer_Master", index=False)
        _format_excel(writer, "Customer_Master", master_df)

    print(f"Wrote customer master: {output_path}")


# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------

def derive_output_paths(input_path: Path) -> Tuple[Path, Path]:
    """Derive output filenames from input filename."""
    stem = input_path.stem  # filename without extension
    parent = input_path.parent

    prepared_path = parent / f"{stem}.prepared.xlsx"
    customers_path = parent / f"{stem}.customers.xlsx"

    return prepared_path, customers_path


def auto_merge_high_confidence(master_df: pd.DataFrame) -> pd.DataFrame:
    """Auto-apply HIGH confidence consolidation suggestions."""
    master_df = master_df.copy()

    for idx, row in master_df.iterrows():
        if row.get("confidence") == "HIGH" and pd.notna(row.get("suggested_consolidation")):
            master_df.at[idx, "customer_master"] = row["suggested_consolidation"]

    return master_df


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="Prepare customer revenue data for analysis."
    )
    parser.add_argument(
        "--input", required=True,
        help="Path to the source Excel file (month-wise customer revenue)."
    )
    parser.add_argument(
        "--output",
        help="Path for output file (derived from input if not specified)."
    )
    parser.add_argument(
        "--master",
        help="Path to existing customer master file to carry forward mappings."
    )
    parser.add_argument(
        "--customers-only", action="store_true",
        help="Only generate customer master file (skip prepared data)."
    )
    parser.add_argument(
        "--data-only", action="store_true",
        help="Only generate prepared data file (skip customer master)."
    )
    parser.add_argument(
        "--merge", action="store_true",
        help="Auto-apply HIGH confidence consolidation suggestions."
    )
    args = parser.parse_args(argv)

    if args.customers_only and args.data_only:
        print("Error: Cannot specify both --customers-only and --data-only")
        return 1

    input_path = Path(args.input).expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    # Determine output path(s)
    if args.output:
        output_path = Path(args.output).expanduser().resolve()
        if args.customers_only:
            customers_path = output_path
            prepared_path = None
        elif args.data_only:
            prepared_path = output_path
            customers_path = None
        else:
            # If --output specified without mode, treat as prepared path
            prepared_path = output_path
            customers_path = output_path.parent / f"{output_path.stem.replace('.prepared', '')}.customers.xlsx"
    else:
        prepared_path, customers_path = derive_output_paths(input_path)
        if args.customers_only:
            prepared_path = None
        elif args.data_only:
            customers_path = None

    # Load existing master if provided
    existing_master = None
    if args.master:
        master_path = Path(args.master).expanduser().resolve()
        if not master_path.exists():
            raise FileNotFoundError(f"Master file not found: {master_path}")
        existing_master = pd.read_excel(master_path, engine="openpyxl")
        print(f"Loaded existing master: {master_path}")

    # Load raw data
    print(f"Loading: {input_path}")
    df_long = load_monthly_customer_revenue(input_path)
    print(f"  Loaded {len(df_long)} revenue records, {df_long['customer'].nunique()} unique customers")

    # Generate customer master if needed
    master_df = None
    if not args.data_only:
        master_df = generate_customer_master(df_long, existing_master)

        # Auto-merge if requested
        if args.merge:
            master_df = auto_merge_high_confidence(master_df)
            print("  Applied HIGH confidence consolidations")

    # Apply mappings to consolidate customers (use existing master or newly generated one)
    mapping_source = existing_master if existing_master is not None else master_df
    if mapping_source is not None and not args.customers_only:
        df_long = apply_customer_master(df_long, mapping_source)
        print(f"  Applied mappings, now {df_long['customer'].nunique()} unique customers")

    # Write outputs
    if prepared_path:
        write_prepared_file(df_long, prepared_path)

    if customers_path and master_df is not None:
        write_customer_master_file(master_df, customers_path)

        # Summary
        new_customers = master_df[master_df["is_new"] == True]["customer_normalized"].nunique()
        if new_customers > 0:
            print(f"\nNote: {new_customers} new customer(s) added to master for review")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
