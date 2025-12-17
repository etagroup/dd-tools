#!/usr/bin/env python3
"""
Generate customer analytics workbook from prepared revenue data.

Reads a prepared Excel file (output from prepare.py) and generates:
- Customer summary with metrics and segmentation
- Rolling concentration analysis
- Top customer tables

Usage:
  python analytics.py --input "data.prepared.xlsx" --output "customer_analytics.xlsx"
"""

from __future__ import annotations

import argparse
import zipfile
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd

# Fixed core.xml for reproducible Excel output
_FIXED_CORE_XML = b'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"><dc:creator>openpyxl</dc:creator><dcterms:created xsi:type="dcterms:W3CDTF">2020-01-01T00:00:00Z</dcterms:created><dcterms:modified xsi:type="dcterms:W3CDTF">2020-01-01T00:00:00Z</dcterms:modified></cp:coreProperties>'''


def _make_xlsx_deterministic(file_path: Path) -> None:
    """Rewrite xlsx file with fixed metadata for reproducible output."""
    with open(file_path, 'rb') as f:
        data = BytesIO(f.read())

    with zipfile.ZipFile(data, 'r') as zf_in:
        file_list = zf_in.namelist()
        contents = {name: zf_in.read(name) for name in file_list}

    # Replace core.xml with fixed version
    contents['docProps/core.xml'] = _FIXED_CORE_XML

    # Fixed timestamp for zip entries (2020-01-01 00:00:00)
    fixed_date_time = (2020, 1, 1, 0, 0, 0)

    with zipfile.ZipFile(file_path, 'w', zipfile.ZIP_DEFLATED) as zf_out:
        for name in sorted(file_list):  # Sort for deterministic order
            info = zipfile.ZipInfo(name, date_time=fixed_date_time)
            info.compress_type = zipfile.ZIP_DEFLATED
            zf_out.writestr(info, contents[name])


# -----------------------------------------------------------------------------
# Rolling Window Calculations
# -----------------------------------------------------------------------------

def _rolling_matrix(monthly_matrix: pd.DataFrame, window_months: int) -> pd.DataFrame:
    """
    Rolling trailing sum by customer.

    Uses min_periods=window_months so early periods are NaN (not partial windows).
    """
    return monthly_matrix.rolling(window=window_months, min_periods=window_months).sum()


def _topk_shares(values: np.ndarray, total: float, ks: Sequence[int]) -> Dict[int, float]:
    """
    Compute top-k share(s) for a 1D array of customer values.
    Returns {k: share}.
    """
    out: Dict[int, float] = {k: np.nan for k in ks}
    if total <= 0 or not np.isfinite(total):
        return out

    v = np.nan_to_num(values, nan=0.0)
    v_sorted = np.sort(v)[::-1]
    cumsum = np.cumsum(v_sorted)
    for k in ks:
        if k <= 0:
            out[k] = np.nan
        else:
            kk = min(k, len(v_sorted))
            out[k] = float(cumsum[kk - 1] / total)
    return out


def build_rolling_concentration(
    monthly_matrix: pd.DataFrame,
    windows: Sequence[int] = (12, 24, 36),
    top_ks: Sequence[int] = (1, 5, 10),
) -> pd.DataFrame:
    """Build a time series table of rolling concentration stats."""
    dates = monthly_matrix.index
    total_monthly = monthly_matrix.sum(axis=1)

    out = pd.DataFrame({"date": dates, "total_monthly_revenue": total_monthly.values})

    for w in windows:
        roll = _rolling_matrix(monthly_matrix, w)
        totals = roll.sum(axis=1)
        top1_customer = []
        top1_revenue = []
        top_shares = {k: [] for k in top_ks}

        for d in dates:
            row = roll.loc[d].to_numpy()
            tot = totals.loc[d]
            shares = _topk_shares(row, float(tot) if pd.notna(tot) else float("nan"), top_ks)
            for k, v in shares.items():
                top_shares[k].append(v)

            if pd.isna(tot) or tot == 0:
                top1_customer.append(np.nan)
                top1_revenue.append(np.nan)
            else:
                s = roll.loc[d]
                idx = s.idxmax()
                top1_customer.append(idx)
                top1_revenue.append(float(s.loc[idx]))

        out[f"total_trailing_{w}m"] = totals.values
        out[f"top1_customer_trailing_{w}m"] = top1_customer
        out[f"top1_revenue_trailing_{w}m"] = top1_revenue
        for k in top_ks:
            out[f"top{k}_share_trailing_{w}m"] = top_shares[k]

    return out


# -----------------------------------------------------------------------------
# Customer-Level Analytics
# -----------------------------------------------------------------------------

def _compute_gap_stats(active_dates: List[pd.Timestamp]) -> Tuple[float, float, int]:
    """
    Given sorted active dates (month starts), return:
      (avg_gap_months, max_gap_months, reactivations)
    """
    if len(active_dates) <= 1:
        return (0.0, 0.0, 0)

    gaps: List[int] = []
    reactivations = 0
    for prev, cur in zip(active_dates[:-1], active_dates[1:]):
        diff_m = (cur.year - prev.year) * 12 + (cur.month - prev.month)
        gap = max(diff_m - 1, 0)
        gaps.append(gap)
        if gap >= 1:
            reactivations += 1

    return (float(np.mean(gaps)), float(np.max(gaps)), int(reactivations))


def build_customer_summary(df_long: pd.DataFrame) -> pd.DataFrame:
    """Build a customer-level summary table for repeat and lumpy behavior."""
    df_long = df_long.sort_values(["customer", "date"]).copy()

    monthly_matrix = (
        df_long.pivot_table(index="date", columns="customer", values="revenue", aggfunc="sum", fill_value=0.0)
        .sort_index()
    )

    lifetime_rev = monthly_matrix.sum(axis=0)

    end_date = monthly_matrix.index.max()
    ttm = _rolling_matrix(monthly_matrix, 12)
    t24 = _rolling_matrix(monthly_matrix, 24)
    t36 = _rolling_matrix(monthly_matrix, 36)

    ttm_last = ttm.loc[end_date] if end_date in ttm.index else pd.Series(dtype=float)
    t24_last = t24.loc[end_date] if end_date in t24.index else pd.Series(dtype=float)
    t36_last = t36.loc[end_date] if end_date in t36.index else pd.Series(dtype=float)

    nonzero = df_long[df_long["revenue"].abs() > 1e-9].copy()
    first_purchase = nonzero.groupby("customer")["date"].min()
    last_purchase = nonzero.groupby("customer")["date"].max()
    active_months = nonzero.groupby("customer")["date"].nunique()
    active_years = nonzero.groupby("customer")["year"].nunique()

    nonzero["quarter"] = nonzero["date"].dt.to_period("Q").astype(str)
    active_quarters = nonzero.groupby("customer")["quarter"].nunique()

    tenure_months = (
        (last_purchase.dt.year - first_purchase.dt.year) * 12
        + (last_purchase.dt.month - first_purchase.dt.month)
        + 1
    )

    activity_ratio = active_months / tenure_months

    gap_avg = {}
    gap_max = {}
    reactivations = {}

    for cust, grp in nonzero.groupby("customer"):
        dates = sorted(grp["date"].unique())
        dates = [pd.Timestamp(d) for d in dates]
        avg_gap, max_gap, react = _compute_gap_stats(dates)
        gap_avg[cust] = avg_gap
        gap_max[cust] = max_gap
        reactivations[cust] = react

    gap_avg_s = pd.Series(gap_avg)
    gap_max_s = pd.Series(gap_max)
    reactivations_s = pd.Series(reactivations)

    is_repeat = (active_years >= 2)

    valid_ttm_dates = ttm.sum(axis=1).dropna().index
    ttm_valid = ttm.loc[valid_ttm_dates].copy()
    ttm_totals = ttm_valid.sum(axis=1)
    share = ttm_valid.div(ttm_totals, axis=0)
    peak_ttm_share = share.max(axis=0)

    top10_counts = pd.Series(0, index=ttm_valid.columns, dtype=int)
    for d in valid_ttm_dates:
        top10 = ttm_valid.loc[d].nlargest(10).index
        top10_counts.loc[top10] += 1
    top10_persistence = top10_counts / max(len(valid_ttm_dates), 1)

    is_repeat_reindexed = is_repeat.reindex(lifetime_rev.index)
    is_repeat_reindexed = is_repeat_reindexed.where(is_repeat_reindexed.notna(), False)

    summary = pd.DataFrame({
        "customer": lifetime_rev.index,
        "lifetime_revenue": lifetime_rev.values,
        "ttm_revenue_last": ttm_last.reindex(lifetime_rev.index).values,
        "t24m_revenue_last": t24_last.reindex(lifetime_rev.index).values,
        "t36m_revenue_last": t36_last.reindex(lifetime_rev.index).values,
        "first_purchase_month": first_purchase.reindex(lifetime_rev.index).values,
        "last_purchase_month": last_purchase.reindex(lifetime_rev.index).values,
        "active_months": active_months.reindex(lifetime_rev.index).fillna(0).astype(int).values,
        "active_quarters": active_quarters.reindex(lifetime_rev.index).fillna(0).astype(int).values,
        "active_years": active_years.reindex(lifetime_rev.index).fillna(0).astype(int).values,
        "tenure_months": tenure_months.reindex(lifetime_rev.index).fillna(0).astype(int).values,
        "activity_ratio": activity_ratio.reindex(lifetime_rev.index).fillna(0.0).values,
        "avg_gap_months": gap_avg_s.reindex(lifetime_rev.index).fillna(0.0).values,
        "max_gap_months": gap_max_s.reindex(lifetime_rev.index).fillna(0.0).values,
        "reactivations": reactivations_s.reindex(lifetime_rev.index).fillna(0).astype(int).values,
        "is_repeat_customer": is_repeat_reindexed.values,
        "peak_ttm_share": peak_ttm_share.reindex(lifetime_rev.index).values,
        "top10_ttm_persistence": top10_persistence.reindex(lifetime_rev.index).values,
    })

    summary = summary.sort_values("lifetime_revenue", ascending=False).reset_index(drop=True)
    return summary


# -----------------------------------------------------------------------------
# Excel Output
# -----------------------------------------------------------------------------

def _format_excel(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame, freeze_panes: Tuple[int, int] = (1, 0)) -> None:
    """Apply lightweight formatting via openpyxl."""
    from openpyxl.utils import get_column_letter

    ws = writer.sheets[sheet_name]
    ws.freeze_panes = ws.cell(row=freeze_panes[0] + 1, column=freeze_panes[1] + 1)
    ws.auto_filter.ref = ws.dimensions

    for col_idx, col in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        col_name = str(col).lower()

        try:
            max_len = max([len(str(col))] + [len(str(x)) for x in df[col].head(200).tolist()])
        except Exception:
            max_len = len(str(col))
        width = min(max(10, max_len + 2), 45)
        ws.column_dimensions[col_letter].width = width

        num_format = None
        if 'share' in col_name or 'ratio' in col_name or 'persistence' in col_name:
            num_format = '0%'
        elif col_name == 'date' or col_name.endswith('_month') or col_name.endswith('_date'):
            num_format = 'yyyy-mm-dd'
        elif 'gap' in col_name:
            num_format = '0.0'
        elif 'revenue' in col_name or 'total_trailing' in col_name or col_name.startswith('total_'):
            num_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'

        if num_format:
            for row in range(2, ws.max_row + 1):
                ws[f'{col_letter}{row}'].number_format = num_format


def _add_customer_segmentation_columns(writer: pd.ExcelWriter, data_end: pd.Timestamp) -> None:
    """Add formula-based segmentation columns to Customer_Summary sheet."""
    from openpyxl.utils import get_column_letter

    ws = writer.sheets["Customer_Summary"]

    col_last_purchase = 7  # last_purchase_month is column G
    col_ttm = 3  # ttm_revenue_last is column C
    col_lifetime = 2  # lifetime_revenue is column B
    col_tenure = 11  # tenure_months is column K
    col_peak_share = 17  # peak_ttm_share is column Q

    num_rows = ws.max_row
    next_col = ws.max_column + 1

    headers = [
        "months_since_last_purchase",
        "status",
        "is_high_value",
        "segment"
    ]

    for i, header in enumerate(headers):
        col_letter = get_column_letter(next_col + i)
        ws[f"{col_letter}1"] = header
        ws.column_dimensions[col_letter].width = 20

    for row in range(2, num_rows + 1):
        col_idx = next_col

        last_purchase_cell = f"{get_column_letter(col_last_purchase)}{row}"
        formula1 = f'=DATEDIF({last_purchase_cell},DATE({data_end.year},{data_end.month},1),"M")'
        ws[f"{get_column_letter(col_idx)}{row}"] = formula1

        col_idx += 1

        months_since_col = get_column_letter(col_idx - 1)
        formula2 = f'=IF({months_since_col}{row}<=6,"Active",IF({months_since_col}{row}<=18,"Inactive","Churned"))'
        ws[f"{get_column_letter(col_idx)}{row}"] = formula2

        col_idx += 1

        lifetime_col = get_column_letter(col_lifetime)
        peak_share_col = get_column_letter(col_peak_share)
        formula3 = f'=OR({lifetime_col}{row}>=1000000,{peak_share_col}{row}>=0.02)'
        ws[f"{get_column_letter(col_idx)}{row}"] = formula3

        col_idx += 1

        status_col = get_column_letter(next_col + 1)
        ttm_col = get_column_letter(col_ttm)
        tenure_col = get_column_letter(col_tenure)
        formula4 = f'=IF({status_col}{row}="Active",IF(AND({ttm_col}{row}>=2000000,OR({tenure_col}{row}>=36,{lifetime_col}{row}>=5000000)),"Strategic",IF({ttm_col}{row}>=1000000,"High Value","Mid Value")),"")'
        ws[f"{get_column_letter(col_idx)}{row}"] = formula4

    ws.auto_filter.ref = ws.dimensions


def write_analytics_workbook(
    df_long: pd.DataFrame,
    output_path: Path,
    data_start: pd.Timestamp,
    data_end: pd.Timestamp
) -> None:
    """Generate the analytics workbook."""
    monthly_matrix = (
        df_long.pivot_table(index="date", columns="customer", values="revenue", aggfunc="sum", fill_value=0.0)
        .sort_index()
    )

    rolling_conc = build_rolling_concentration(monthly_matrix, windows=(12, 24, 36), top_ks=(1, 5, 10))
    cust_summary = build_customer_summary(df_long)

    top25_lifetime = cust_summary.sort_values("lifetime_revenue", ascending=False).head(25)
    top25_ttm = cust_summary.sort_values("ttm_revenue_last", ascending=False).head(25)
    top25_peak = cust_summary.sort_values("peak_ttm_share", ascending=False).head(25)

    # Metadata for reports (no timestamp for reproducible output)
    metadata = pd.DataFrame({
        "property": ["data_start_date", "data_end_date"],
        "value": [
            data_start.strftime("%Y-%m-%d"),
            data_end.strftime("%Y-%m-%d"),
        ]
    })

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Metadata first (for easy access by reports)
        metadata.to_excel(writer, sheet_name="Metadata", index=False)

        # Customer summary
        cust_summary.to_excel(writer, sheet_name="Customer_Summary", index=False)
        _format_excel(writer, "Customer_Summary", cust_summary)
        _add_customer_segmentation_columns(writer, data_end)

        # Monthly matrix
        monthly_matrix.to_excel(writer, sheet_name="Monthly_Matrix")
        _format_excel(writer, "Monthly_Matrix", monthly_matrix.reset_index())

        # Rolling concentration
        rolling_conc.to_excel(writer, sheet_name="Rolling_Concentration", index=False)
        _format_excel(writer, "Rolling_Concentration", rolling_conc)

        # Top tables
        top25_lifetime.to_excel(writer, sheet_name="Top25_Lifetime", index=False)
        _format_excel(writer, "Top25_Lifetime", top25_lifetime)

        top25_ttm.to_excel(writer, sheet_name="Top25_TTM", index=False)
        _format_excel(writer, "Top25_TTM", top25_ttm)

        top25_peak.to_excel(writer, sheet_name="Top25_PeakTTMShare", index=False)
        _format_excel(writer, "Top25_PeakTTMShare", top25_peak)

    _make_xlsx_deterministic(output_path)
    print(f"Wrote analytics workbook: {output_path}")


# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------

def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="Generate customer analytics workbook from prepared data."
    )
    parser.add_argument(
        "--input", required=True,
        help="Path to prepared Excel file (output from prepare.py)."
    )
    parser.add_argument(
        "--output", required=True,
        help="Path to write the analytics Excel workbook."
    )
    args = parser.parse_args(argv)

    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    # Load prepared data
    print(f"Loading: {input_path}")
    df_long = pd.read_excel(input_path, sheet_name="Revenue_Detail")
    df_long["date"] = pd.to_datetime(df_long["date"])

    # Get date range
    data_start = df_long["date"].min()
    data_end = df_long["date"].max()

    print(f"  Data range: {data_start.strftime('%b %Y')} - {data_end.strftime('%b %Y')}")
    print(f"  Customers: {df_long['customer'].nunique()}")
    print(f"  Records: {len(df_long)}")

    # Generate analytics workbook
    write_analytics_workbook(df_long, output_path, data_start, data_end)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
